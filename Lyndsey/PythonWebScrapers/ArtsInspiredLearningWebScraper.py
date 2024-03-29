#!/usr/bin/python

# ChromeDriver woes: make sure it's up to date https://stackoverflow.com/questions/21166408/how-do-i-confirm-im-using-the-right-chromedriver
# or just rename the .exe at C:\Windows and it maybe just works?

from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import datetime
from openpyxl import Workbook
import re
import os

# IsDebugMode = True
IsDebugMode = False

LoadAllPrograms = True
# LoadAllPrograms = False

SingleProgramUrl = ""
# SingleProgramUrl = "https://arts-inspiredlearning.org/programs/three-for-one-exploring-geometry-with-a-transforming-origami-puppet/"

rootUrl = "https://arts-inspiredlearning.org/programs/"
# rootUrl = "https://arts-inspiredlearning.org/virtual-programming/" # no need to load more, different home page (but same "program" pages). But also turns out I didn't need to do this, the ones with content were duplicates of the other page

justProgramDetails = True
# justProgramDetails = False
sqaurespaceProgramRootUrl = "https://center-for-arts.squarespace.com/arts-programs/p/"

proramRootUrl = "https://arts-inspiredlearning.org/programs"
scrapedResultsFilename = "ArtsInspiredLearningScraped.xlsx"
logFilePath = "C:\src\personalthings\Lyndsey\PythonWebScrapers\LogFile.log"
AllProgramTags = []

if os.path.exists(logFilePath):
  os.remove(logFilePath)

class Detail:
    header: ""
    values: [] # should be a single item (maybe comma separated, but still a single "item") except for the last section, which is a bulleted list

    def __init__(self, header, values):
        self.header = header
        self.values = values
    
    def prettyPrint(self):
        joined = " | ".join(self.values)
        return f"Header: {self.header} - Values: {joined}"
    
    def asHtml(self):
        items = ""
        if (len(self.values) == 1):
            items = f"<p2>{self.values[0]}</p2>"
        elif (len(self.values) > 1):
            items = "<p2><ul><li>" + "</li><li>".join(self.values) + "</li></ul></p2>"
        return f"<h4>{self.header}</h4>{items}<hr align=\"left\" />"

class ProgramDetails:
    routePath = ""
    name = ""
    rawDescription = ""
    formattedDescription = ""
    artists = []
    tags = "" # called program details in the website, Lyndsey wants them as a comma-separated list
    programDetailsHtml = "" # specially formatted "tags" to be copy/pasted into Squarespace

    def prettyPrint(self):
        return f"Name: {self.name}\nRoute Path: {self.routePath}\nRaw Description: {self.rawDescription}\nFormatted Description: {self.formattedDescription}\nArtists: {self.artists}\nTags: {self.tags}\nProgram Details HTML: {self.programDetailsHtml}"

def Log(asDebug, message):
    messagePlusNewline = str(message)+"\n"
    if not asDebug:
        print(message)
        f = open(logFilePath, "ab")
        f.write(messagePlusNewline.encode("UTF-8"))
        f.close()
    if asDebug and IsDebugMode:
        print(message)
        f = open(logFilePath, "ab")
        f.write(messagePlusNewline.encode("UTF-8"))
        f.close()

def dedupeList(items):
    return list( dict.fromkeys(items) )

def collectLinksToPrograms(browser):
    parentElements = browser.find_elements(By.CLASS_NAME, "bbq-content")
    if len(parentElements) > 0:
        containerElement = parentElements[0]
    else:
        parentElements = browser.find_elements(By.CLASS_NAME, "entry-content")
        print(f"{len(parentElements)} parent elements")
        if len(parentElements) > 0:
            containerElement = parentElements[0]
        else:
            Log(False, "Couldn't find parent element to acquire program links!")
            return []
    
    elements = containerElement.find_elements(By.TAG_NAME, "a")
    Log(True, f"We have {len(elements)} 'a' elements to work with")
    links = []
    for e in elements:
        href = e.get_attribute("href")
        if href is not None and proramRootUrl in href.lower():
            links.append(href)
    Log(True, f"We have {len(links)} program links to work with")
    return links

def loadMore(browser, previousLinkCount):
    loadMoreDiv = browser.find_element(By.CLASS_NAME, "fl-builder-pagination-load-more")
    loadMoreButton = loadMoreDiv.find_element(By.TAG_NAME, "a")
    if loadMoreButton is not None:
        loadMoreButton.click()
        time.sleep(2)
        linkCount = len(collectLinksToPrograms(browser))
        if linkCount != previousLinkCount:
            Log(True, f"We had {previousLinkCount}, now we have {linkCount}")
            loadMore(browser, linkCount)
    
def getDataForProgram(browser, linkToProgram):
    browser.get(linkToProgram)
    primarySectionElement = browser.find_element(By.ID, "primary")
    programName = primarySectionElement.find_element(By.TAG_NAME, "h1").text
    Log(True, programName)
    if programName.lower() == "not found":
        Log(False, f"This appears to be a 'Not Found' page, gonna skip it - {linkToProgram}")
        return None

    descriptionElement = primarySectionElement.find_element(By.CLASS_NAME, "entry-content")
    rawDescription = descriptionElement.text
    Log(True, rawDescription)
    formattedDescription = descriptionElement.get_attribute("innerHTML")
    Log(True, formattedDescription)

    secondarySectionElement = browser.find_element(By.ID, "secondary")
    programArtistsContainerElements = secondarySectionElement.find_elements(By.CLASS_NAME, "sidebar-program-artists")
    artists = []
    if len(programArtistsContainerElements) > 0:
        programArtistsContainerElement = programArtistsContainerElements[0]
        programArtistElements = programArtistsContainerElement.find_elements(By.TAG_NAME, "a")
        for e in programArtistElements:
            artists.append(e.text)
        artists = dedupeList(artists)
    Log(True, artists)

    programDetailsContainerElement = secondarySectionElement.find_element(By.CLASS_NAME, "sidebar-program-details")
    items = programDetailsContainerElement.find_elements(By.TAG_NAME, "li")
    Log(True, f"I see {len(items)} items in program details")
    tags = []
    details = []
    for e in items:
        tagText = e.text
        splitTags = []
        Log(True, f"Raw tag text: {tagText}")
        detailHeaderElements = e.find_elements(By.XPATH, "h4 | h6")
        if len(detailHeaderElements) > 0:
            detailHeaderText = detailHeaderElements[0].text
            Log(True, f"detailHeaderText: {detailHeaderText}")
            tagText = tagText.replace(detailHeaderText, "")
        splitTags = tagText.strip().split("\n") # because the last section is a li within a li so it ends up being duplicated, once with newlines and once as individual tags. Splitting on newline lets my dedupe logic work better
        while("" in splitTags):
            splitTags.remove("")
        tags.extend(splitTags)
        while("In-person Option" in splitTags):
            splitTags.remove("In-person Option")
        l = list(filter(lambda x: x.header == "PROGRAM AVAILABLE AS:", details))
        if (len(l) == 0):
            # since that last section comes through more than once, we only want to grab it the first time when it has all of its pieces. The subsequent times are just once per nested li element
            details.append(Detail(detailHeaderText, splitTags))
    while("" in tags):
        tags.remove("")
    tags = dedupeList(tags)
    Log(True, tags)
    AllProgramTags.extend(tags)

    Log(True, "DETAILS:")
    for d in details:
        Log(True, d.asHtml())
    
    programDetailsHtml = "<h3>Program Details</h3>"
    for d in filter(lambda x: len(x.values) > 0, details):
        programDetailsHtml += d.asHtml()
    programDetailsHtml = programDetailsHtml[:-19] # no time for good code! Gotta remove the trailing <hr>

    program = ProgramDetails()
    program.routePath = linkToProgram.replace(rootUrl, "").rstrip("/")
    program.name = programName
    program.rawDescription = rawDescription
    program.formattedDescription = formattedDescription
    program.artists = artists
    program.tags = ", ".join(tags)
    program.programDetailsHtml = programDetailsHtml
    print("\n")
    Log(False, program.prettyPrint())

    return program

def writeHeadersToWorkbook(sheet):
    if (justProgramDetails):
        sheet["A1"] = "Squarespace Product URL"
        sheet["B1"] = "ProgramDetails HTML"
        sheet["C1"] = "Artists"
    else:
        sheet["A1"] = "Product ID [Non Editable]"
        sheet["B1"] = "Variant ID [Non Editable]"
        sheet["C1"] = "Product Type [Non Editable]"
        sheet["D1"] = "Product Page"
        sheet["E1"] = "Product URL"
        sheet["F1"] = "Title"
        sheet["G1"] = "Description"
        sheet["H1"] = "SKU"
        sheet["I1"] = "Option Name 1"
        sheet["J1"] = "Option Value 1"
        sheet["K1"] = "Option Name 2"
        sheet["L1"] = "Option Value 2"
        sheet["M1"] = "Option Name 3"
        sheet["N1"] = "Option Value 3"
        sheet["O1"] = "Option Name 4"
        sheet["P1"] = "Option Value 4"
        sheet["Q1"] = "Option Name 5"
        sheet["R1"] = "Option Value 5"
        sheet["S1"] = "Option Name 6"
        sheet["T1"] = "Option Value 6"
        sheet["U1"] = "Price"
        sheet["V1"] = "Sale Price"
        sheet["W1"] = "On Sale"
        sheet["X1"] = "Stock"
        sheet["Y1"] = "Categories"
        sheet["Z1"] = "Tags"
        sheet["AA1"] = "Weight"
        sheet["AB1"] = "Length"
        sheet["AC1"] = "Width"
        sheet["AD1"] = "Height"
        sheet["AE1"] = "Visible"
        sheet["AF1"] = "Hosted Image URLs"
    
def writeProgramToWorkbook(sheet, index, program):
    if (justProgramDetails):
        sheet["A"+str(index)] = f'{sqaurespaceProgramRootUrl}{program.routePath}'
        sheet["B"+str(index)] = program.programDetailsHtml
        sheet["C"+str(index)] = ", ".join(program.artists)
    else:
        sheet["A"+str(index)] = ""
        sheet["B"+str(index)] = ""
        sheet["C"+str(index)] = "SERVICE"
        sheet["D"+str(index)] = "arts-programs"
        sheet["E"+str(index)] = program.routePath
        sheet["F"+str(index)] = program.name
        sheet["G"+str(index)] = program.formattedDescription
        sheet["H"+str(index)] = "SQ4218371"
        sheet["I"+str(index)] = ""
        sheet["J"+str(index)] = ""
        sheet["K"+str(index)] = ""
        sheet["L"+str(index)] = ""
        sheet["M"+str(index)] = ""
        sheet["N"+str(index)] = ""
        sheet["O"+str(index)] = ""
        sheet["P"+str(index)] = ""
        sheet["Q"+str(index)] = ""
        sheet["R"+str(index)] = ""
        sheet["S"+str(index)] = ""
        sheet["T"+str(index)] = ""
        sheet["U"+str(index)] = "0"
        sheet["V"+str(index)] = "0"
        sheet["W"+str(index)] = "No"
        sheet["X"+str(index)] = "Unlimited"
        sheet["Y"+str(index)] = program.tags
        sheet["Z"+str(index)] = ""
        sheet["AA"+str(index)] = "0"
        sheet["AB"+str(index)] = "0"
        sheet["AC"+str(index)] = "0"
        sheet["AD"+str(index)] = "0"
        sheet["AE"+str(index)] = "Yes"
        sheet["AF"+str(index)] = ""

Log(False, "Start time: " + str(datetime.datetime.now()))

browser = webdriver.Chrome()
browser.get(rootUrl)

if LoadAllPrograms and len(SingleProgramUrl) == 0:
    loadMore(browser, 0)

allProgramLinks = [SingleProgramUrl] if len(SingleProgramUrl) > 0 else collectLinksToPrograms(browser)
Log(False, allProgramLinks)

workbook = Workbook()
sheet = workbook.active
writeHeadersToWorkbook(sheet)    

programIndex = 1
workbookIndex = 2
for link in allProgramLinks:
    Log(False, f'{str(programIndex)}/{str(len(allProgramLinks))}')
    program = getDataForProgram(browser, link)
    if program is not None:
        writeProgramToWorkbook(sheet, workbookIndex, program)
        workbookIndex += 1
    programIndex += 1

workbook.save(filename=scrapedResultsFilename)

browser.quit()

AllProgramTags = dedupeList(AllProgramTags)
print(AllProgramTags)

Log(False, "End time: " + str(datetime.datetime.now()))